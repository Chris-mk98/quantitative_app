import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# --- 상수 및 기본 설정 ---
BASE_ORDERED_COLUMNS_PREFIX = [
    "BvD ID number", "Company name Latin alphabet", "Consolidation code", "SH - BvD Independence Indicator", "Listing status", "US SIC, primary code(s)",
    "US SIC, primary code(s) - description", "Country", "City\nLatin Alphabet", "Website address", "Date of incorporation", "Full overview", "Status", "Main activity",
    "Primary business line", "Main products and services"
]

BASE_ORDERED_COLUMNS_PREFIX_KOR = {
    "BcD ID" : "BvD ID number", "회사명" : "Company name Latin alphabet", "Consolidation code" : "Consolidation code", "독립성 지표" : "SH - BvD Independence Indicator", "상장여부" : "Listing status", "US SIC 코드" : "US SIC, primary code(s)",
    "US SIC 코드 설명" : "US SIC, primary code(s) - description", "국가" : "Country", "도시" : "City\nLatin Alphabet", "웹사이트" :  "Website address", "설립일" : "Date of incorporation", "회사개요" : "Full overview", "스테이터스" : "Status", "주요활동(기능)" : "Main activity",
    "주요사업" : "Primary business line", "주요 상품 및 서비스" : "Main products and services"
}

BASE_ORDERED_COLUMNS_YEARLY = [
    "Number of employees\n",
    "Audit status\n",
    "Operating revenue (Turnover)\nth USD ",
    "Sales\nth USD ",
    "Costs of goods sold\nth USD ",
    "Gross profit\nth USD ",
    "Other operating expense (income)\nth USD ",
    "Operating profit (loss) [EBIT]\nth USD ",
    "Research & Development expenses\nth USD "
]

BASE_ORDERED_COLUMNS_YEARLY_KOR = {
    "직원수" : "Number of employees\n",
    "감사의견" : "Audit status\n",
    "매출액(Turnover)" : "Operating revenue (Turnover)\nth USD ",
    "매출액(Sales)" : "Sales\nth USD ",
    "매출원가" : "Costs of goods sold\nth USD ",
    "매출총이익" : "Gross profit\nth USD ",
    "영업비용" : "Other operating expense (income)\nth USD ",
    "영업이익(손실)" : "Operating profit (loss) [EBIT]\nth USD ",
    "연구개발비" : "Research & Development expenses\nth USD "
}

BASE_ORDERED_COLUMNS_ASSET_YEARLY = [
    "Total assets\nth USD ",
    "Debtors\nth USD ",
    "Creditors\nth USD ",
    "Stock\nth USD ",
    "Intangible assets\nth USD ",
    "Tangible fixed assets\nth USD "
]

BASE_ORDERED_COLUMNS_ASSET_YEARLY_KOR = {
    "총자산" : "Total assets\nth USD ",
    "매입채무" : "Debtors\nth USD ",
    "매출채권" : "Creditors\nth USD ",
    "재고자산" : "Stock\nth USD ",
    "무형자산" : "Intangible assets\nth USD ",
    "유형자산" : "Tangible fixed assets\nth USD " 
}

COLOR_CODES = {
    "orange": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    "green": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    "yellow": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
    "gray": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
}

THIN_BORDER = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

BOLD_FONT = Font(bold=True)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')

class CriteriaFormulaGenerator:
    """양적기준 수식 생성 클래스"""
    
    def __init__(self, analysis_instance):
        """
        Analysis 객체를 받아서 모든 정보를 자동으로 가져옴
        
        Parameters:
        - analysis_instance: Analysis 클래스의 인스턴스
        """
        self.analysis = analysis_instance
        self.start_year = analysis_instance.start_year
        self.end_year = analysis_instance.end_year
        self.num_years = analysis_instance.num_years
        self.raw_col_alphabet = analysis_instance.raw_col_alphabet
        self.raw_col_number = analysis_instance.raw_col_number
        self.wa3_start_col = analysis_instance.wa3_start_col
        self.flow_start_col = analysis_instance.flow_start_col
        self.quantitative_start_row = analysis_instance.quantitative_start_row
        self.quantitative_start_col = analysis_instance.quantitative_start_col
        self.number_of_criteria = analysis_instance.number_of_criteria
        
        # Flow 탭 매핑 (자산/부채 항목들)
        self.flow_mapping = {
            "Debtors": 0,  # 매출채권
            "Creditors": 1,  # 매입채무
            "Stock": 2,  # 재고자산
            "Intangible assets": 3,  # 무형자산
            "Tangible fixed assets": 4,  # 유형자산
            "Total assets": 5  # 총자산
        }
        
        # WA3 탭 매핑
        self.wa3_mapping = {
            "매출액": 0,
            "영업이익": 1,
            "영업비용": 2,
            "재고자산": 3,
            "연구개발비": 4,
            "무형자산": 5,
            "유형자산": 6,
            "총자산": 7,
            "매출원가": 8,
            "종업원수": 9
        }
        
        # 비율 탭 매핑 (WA3 다음에 위치)
        self.ratio_mapping = {
            "연구개발비/매출액": 10,
            "영업비용/매출액": 11,
            "무형자산/총자산": 12,
            "유형자산/총자산": 13,
            "재고자산/총자산": 14,
            "재고자산보유일수": 15
        }
    
    def _get_column_range(self, field_name, row_number):
        """필드명에 따라 컬럼 범위를 반환"""
        # Flow 데이터인 경우
        for flow_key in self.flow_mapping:
            if flow_key.lower() in field_name.lower():
                flow_idx = self.flow_mapping[flow_key]
                num_flow_cols = self.num_years + 1  # 개별년도 + 가중평균
                start_col = self.flow_start_col + flow_idx * num_flow_cols
                # 개별 연도만 (가중평균 제외)
                cols = [get_column_letter(start_col + i) for i in range(self.num_years)]
                return [f"{col}{row_number}" for col in cols]
        
        # Raw 데이터인 경우 (연도별)
        if field_name + str(self.start_year) in self.raw_col_alphabet:
            cols = []
            for year in range(self.start_year, self.end_year + 1):
                col_name = field_name + str(year)
                if col_name in self.raw_col_alphabet:
                    cols.append(f"{self.raw_col_alphabet[col_name]}{row_number}")
            return cols
        
        # 단일 컬럼 (연도 없음)
        if field_name in self.raw_col_alphabet:
            return [f"{self.raw_col_alphabet[field_name]}{row_number}"]
        
        return []
    
    def _get_wa3_column(self, metric_name, row_number):
        """WA3/비율 탭에서 특정 지표의 컬럼 위치 반환"""
        if metric_name in self.wa3_mapping:
            col = get_column_letter(self.wa3_start_col + self.wa3_mapping[metric_name])
            return f"{col}{row_number}"
        if metric_name in self.ratio_mapping:
            col = get_column_letter(self.wa3_start_col + self.ratio_mapping[metric_name])
            return f"{col}{row_number}"
        return None
    
    def _get_criteria_threshold_cell(self, criteria_index):
        """
        특정 기준의 threshold 값이 저장될 셀 위치 반환 (절대참조)
        
        Parameters:
        - criteria_index: 기준 번호 (1부터 시작)
        
        Returns:
        - 예: "$D$8" (기준1의 threshold 셀)
        """
        # 기준 threshold는 quantitative_start_row + 2 행에 위치
        threshold_row = self.quantitative_start_row + 2
        threshold_col = self.quantitative_start_col + criteria_index - 1
        col_letter = get_column_letter(threshold_col)
        return f"${col_letter}${threshold_row}"
    
    def generate_text_criteria(self, field_name, condition_type, value, row_number, include=True):
        """
        텍스트 기준 수식 생성
        
        Parameters:
        - field_name: 필드명 (예: "Audit status\n", "Website address")
        - condition_type: 조건 타입
            - "blank": 공란 체크
            - "not_blank": 비공란 체크
            - "equals": 특정 값과 같음
            - "contains": 특정 단어 포함
            - "all_equals": 모든 연도에서 특정 값과 같음
        - value: 비교값 (예: "Unqualified", "적정")
        - row_number: 행 번호
        - include: True면 조건 충족시 포함(Yes), False면 제외(No)
        """
        result = "Yes" if include else "No"
        opposite = "No" if include else "Yes"
        
        cols = self._get_column_range(field_name, row_number)
        
        if not cols:
            return f'="{opposite}"  # Error: Field not found'
        
        if condition_type == "blank":
            # 공란인 경우
            if len(cols) == 1:
                formula = f'=IF({cols[0]}="","{result}","{opposite}")'
            else:
                # 모든 연도가 공란인 경우
                conditions = ",".join([f'{col}=""' for col in cols])
                formula = f'=IF(AND({conditions}),"{result}","{opposite}")'
        
        elif condition_type == "not_blank":
            # 비공란인 경우 (데이터가 있는 경우)
            if len(cols) == 1:
                formula = f'=IF({cols[0]}<>"","{result}","{opposite}")'
            else:
                # 모든 연도에 데이터가 있는 경우
                conditions = ",".join([f'{col}<>""' for col in cols])
                formula = f'=IF(AND({conditions}),"{result}","{opposite}")'
        
        elif condition_type == "equals":
            # 특정 값과 같은 경우
            if len(cols) == 1:
                formula = f'=IF({cols[0]}="{value}","{result}","{opposite}")'
            else:
                # 한 번이라도 같은 경우
                conditions = ",".join([f'{col}="{value}"' for col in cols])
                formula = f'=IF(OR({conditions}),"{result}","{opposite}")'
        
        elif condition_type == "all_equals":
            # 모든 연도에서 특정 값과 같은 경우
            if len(cols) == 1:
                formula = f'=IF({cols[0]}="{value}","{result}","{opposite}")'
            else:
                # COUNTIF 사용 (예: 감사의견 적정)
                first_col = cols[0].replace(str(row_number), '')
                last_col = cols[-1].replace(str(row_number), '')
                range_str = f"${first_col}${row_number}:${last_col}${row_number}"
                formula = f'=IF(COUNTIF({range_str},"{value}")={len(cols)},"{result}","{opposite}")'
        
        elif condition_type == "contains":
            # 특정 단어를 포함하는 경우
            if len(cols) == 1:
                formula = f'=IF(ISNUMBER(SEARCH("{value}",{cols[0]})),"{result}","{opposite}")'
            else:
                # 한 번이라도 포함하는 경우
                conditions = ",".join([f'ISNUMBER(SEARCH("{value}",{col}))' for col in cols])
                formula = f'=IF(OR({conditions}),"{result}","{opposite}")'
        
        else:
            formula = f'="{opposite}"  # Error: Unknown condition type'
        
        return formula
    
    def generate_numeric_criteria(self, field_name, condition_type, threshold, row_number, 
                                  include=True, use_threshold_cell=False, criteria_index=None,
                                  count_requirement=None):
        """
        숫자 기준 수식 생성
        
        Parameters:
        - field_name: 필드명 (예: "Operating profit (loss) [EBIT]\nth USD ")
        - condition_type: 조건 타입
            - "gt": 초과 (>)
            - "gte": 이상 (>=)
            - "lt": 미만 (<)
            - "lte": 이하 (<=)
            - "eq": 같음 (=)
        - threshold: 기준값 (숫자 또는 셀참조용 인덱스)
        - row_number: 행 번호
        - include: True면 조건 충족시 포함(Yes), False면 제외(No)
        - use_threshold_cell: True면 기준값을 셀 참조로 사용
        - criteria_index: threshold 셀 위치 계산용 기준 인덱스 (use_threshold_cell=True일 때 필요)
        - count_requirement: 조건을 충족해야 하는 횟수
            - None or "all": 모든 연도 충족
            - "any": 한 번이라도 충족
            - int: N회 이상 충족
        """
        result = "Yes" if include else "No"
        opposite = "No" if include else "Yes"
        
        cols = self._get_column_range(field_name, row_number)
        
        if not cols:
            return f'="{opposite}"  # Error: Field not found'
        
        # threshold 값 결정
        if use_threshold_cell and criteria_index is not None:
            threshold_ref = self._get_criteria_threshold_cell(criteria_index)
        else:
            threshold_ref = str(threshold)
        
        # 비교 연산자 매핑
        operators = {
            "gt": ">",
            "gte": ">=",
            "lt": "<",
            "lte": "<=",
            "eq": "="
        }
        
        op = operators.get(condition_type, ">")
        
        # 단일 조건 생성
        def make_condition(col):
            return f"{col}{op}{threshold_ref}"
        
        # 횟수 기준이 없는 경우 (모든 연도 충족)
        if count_requirement is None or count_requirement == "all":
            conditions = ",".join([make_condition(col) for col in cols])
            formula = f'=IFERROR(IF(AND({conditions}),"{result}","{opposite}"),"{opposite}")'
        
        # 한 번이라도 충족
        elif count_requirement == "any":
            conditions = ",".join([make_condition(col) for col in cols])
            formula = f'=IFERROR(IF(OR({conditions}),"{result}","{opposite}"),"{opposite}")'
        
        # 특정 횟수 이상 충족
        elif isinstance(count_requirement, int):
            # COUNTIF 사용하여 조건 충족 횟수 세기
            count_conditions = "+".join([f'IF({make_condition(col)},1,0)' for col in cols])
            formula = f'=IFERROR(IF(({count_conditions})>={count_requirement},"{result}","{opposite}"),"{opposite}")'
        
        else:
            formula = f'="{opposite}"  # Error: Unknown count requirement'
        
        return formula
    
    def generate_ratio_criteria(self, ratio_name, condition_type, threshold, row_number, 
                               include=True, use_threshold_cell=False, criteria_index=None):
        """
        비율 기준 수식 생성 (WA3 탭 데이터 사용)
        
        Parameters:
        - ratio_name: 비율명 (예: "연구개발비/매출액", "무형자산/총자산")
        - condition_type: 조건 타입 ("gt", "gte", "lt", "lte", "eq")
        - threshold: 기준값 (예: 0.03 = 3%)
        - row_number: 행 번호
        - include: True면 조건 충족시 포함(Yes), False면 제외(No)
        - use_threshold_cell: True면 기준값을 셀 참조로 사용
        - criteria_index: threshold 셀 위치 계산용 기준 인덱스
        """
        result = "Yes" if include else "No"
        opposite = "No" if include else "Yes"
        
        col = self._get_wa3_column(ratio_name, row_number)
        
        if not col:
            return f'="{opposite}"  # Error: Ratio not found'
        
        # threshold 값 결정
        if use_threshold_cell and criteria_index is not None:
            threshold_ref = self._get_criteria_threshold_cell(criteria_index)
        else:
            threshold_ref = str(threshold)
        
        operators = {
            "gt": ">",
            "gte": ">=",
            "lt": "<",
            "lte": "<=",
            "eq": "="
        }
        
        op = operators.get(condition_type, ">")
        
        formula = f'=IFERROR(IF({col}{op}{threshold_ref},"{result}","{opposite}"),"{opposite}")'
        
        return formula
    
    def generate_data_availability_criteria(self, field_names, row_number, include=True):
        """
        데이터 가용성 체크 (모든 필드에 숫자 데이터가 있는지 확인)
        
        Parameters:
        - field_names: 확인할 필드명 리스트
        - row_number: 행 번호
        - include: True면 모두 있을 때 포함, False면 제외
        """
        result = "Yes" if include else "No"
        opposite = "No" if include else "Yes"
        
        all_cols = []
        for field_name in field_names:
            cols = self._get_column_range(field_name, row_number)
            all_cols.extend(cols)
        
        if not all_cols:
            return f'="{opposite}"  # Error: No fields found'
        
        # ISNUMBER로 모든 셀이 숫자인지 확인
        conditions = ",".join([f"ISNUMBER({col})" for col in all_cols])
        formula = f'=IF(AND({conditions}),"{result}","{opposite}")'
        
        return formula
    
    def generate_wa3_numeric_criteria(self, ratio_name, condition_type, threshold, row_number, 
                               include=True, use_threshold_cell=False, criteria_index=None):
        """
        비율 기준 수식 생성 (WA3 탭 데이터 사용)
        
        Parameters:
        - ratio_name: 비율명 (예: "연구개발비/매출액", "무형자산/총자산")
        - condition_type: 조건 타입 ("gt", "gte", "lt", "lte", "eq")
        - threshold: 기준값 (예: 0.03 = 3%)
        - row_number: 행 번호
        - include: True면 조건 충족시 포함(Yes), False면 제외(No)
        - use_threshold_cell: True면 기준값을 셀 참조로 사용
        - criteria_index: threshold 셀 위치 계산용 기준 인덱스
        """
        result = "Yes" if include else "No"
        opposite = "No" if include else "Yes"
        
        col = self._get_wa3_column(ratio_name, row_number)
        
        if not col:
            return f'="{opposite}"  # Error: Ratio not found'
        
        # threshold 값 결정
        if use_threshold_cell and criteria_index is not None:
            threshold_ref = self._get_criteria_threshold_cell(criteria_index)
        else:
            threshold_ref = str(threshold)
        
        operators = {
            "gt": ">",
            "gte": ">=",
            "lt": "<",
            "lte": "<=",
            "eq": "="
        }
        
        op = operators.get(condition_type, ">")
        
        formula = f'=IFERROR(IF({col}{op}{threshold_ref},"{result}","{opposite}"),"{opposite}")'
        
        return formula

class Analysis:
    def __init__(self, tested_party="test", start_year=2021, end_year=2023, name="test", number_of_criteria=5, data_path="", criteria_list=None):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.tested_party = tested_party
        self.start_year = start_year
        self.end_year = end_year
        self.name = name
        self.number_of_criteria = number_of_criteria
        self.data_path = data_path
        self.criteria_list = criteria_list if criteria_list else []
        self.color_code = COLOR_CODES
        
        self.num_years = self.end_year - self.start_year + 1
        self.ordered_columns = self._generate_dynamic_ordered_columns()
        
        self.quantitative_start_row = self.number_of_criteria + 20
        self.quantitative_start_col = 4
        self.qualitative_start_row = self.quantitative_start_row
        self.qualitative_start_col = self.quantitative_start_col + self.number_of_criteria + 1
        self.selection_history_start_col = self.qualitative_start_col + len(self._get_qualitative_criteria_keys())
        self.unadjusted_start_col = self.selection_history_start_col + 3
        self.unadjusted_num_cols = (self.num_years + 2) * 3
        self.wa3_start_col = self.unadjusted_start_col + self.unadjusted_num_cols
        self.raw_data_start_col = self.wa3_start_col + len(self._get_wa3_list()) + len(self._get_ratio_tab_list())
        self.flow_start_col = self.raw_data_start_col + len(self.ordered_columns)
        
        self.raw_col_number = {name : self.ordered_columns.index(name)+self.raw_data_start_col for name in self.ordered_columns}
        self.raw_col_alphabet = {name : get_column_letter(self.ordered_columns.index(name)+self.raw_data_start_col) for name in self.ordered_columns}
        
        self.max_formatted_col = 0
        self.max_formatted_row = 0
        
        # CriteriaFormulaGenerator 초기화
        self.formula_generator = CriteriaFormulaGenerator(self)

    def _generate_yearly_columns(self, base_names, include_start_year_minus_one=False):
        yearly_columns = []
        years_to_include = list(range(self.start_year, self.end_year + 1))
        if include_start_year_minus_one:
            years_to_include.insert(0, self.start_year - 1)
        for base_name in base_names:
            for year in years_to_include:
                yearly_columns.append(f"{base_name}{year}")
        return yearly_columns

    def _generate_dynamic_ordered_columns(self):
        dynamic_columns = []
        dynamic_columns.extend(BASE_ORDERED_COLUMNS_PREFIX)
        dynamic_columns.extend(self._generate_yearly_columns(BASE_ORDERED_COLUMNS_YEARLY, include_start_year_minus_one=False))
        dynamic_columns.extend(self._generate_yearly_columns(BASE_ORDERED_COLUMNS_ASSET_YEARLY, include_start_year_minus_one=True))
        return dynamic_columns

    def _get_qualitative_criteria_data(self):
        return {
            "DB\nDescription": "", "기준1": "메인액티비티", "기준2": "기타",
            "기준3": "제품 또는 용역", "기준4": "정보불충분", "기준5": "특이사건",
            "기준6": "기타", "비고": "", "질적통과(1차)\n전기선정": "",
            "질적통과(2차)\n당기선정": ""
        }

    def _get_qualitative_criteria_keys(self):
        return list(self._get_qualitative_criteria_data().keys())

    def _get_unadj_list(self):
        unadj_years = [f'FY{year - 2000}' for year in range(self.start_year, self.end_year + 1)]
        unadj_years.append(f"FY{self.start_year-2000}-{self.end_year-2000}")
        unadj_years.append("Max-Min")
        return unadj_years

    def _get_wa3_list(self):
        return ["매출액", "영업이익", "영업비용", "재고자산", "연구개발비", "무형자산", "유형자산", "총자산", "매출원가", "종업원수"]

    def _get_ratio_tab_list(self):
        return ["연구개발비/매출액", "영업비용/매출액", "무형자산/총자산", "유형자산/총자산", "재고자산/총자산", "재고자산보유일수\n(365/재고자산회전율)"]

    def _apply_common_styles(self, min_row, max_row, min_col, max_col):
        for r_idx in range(min_row, max_row + 1):
            for c_idx in range(min_col, max_col + 1):
                cell = self.ws.cell(row=r_idx, column=c_idx)
                cell.border = THIN_BORDER
                cell.font = BOLD_FONT
                cell.alignment = CENTER_ALIGN
        self.max_formatted_col = max(self.max_formatted_col, max_col)
        self.max_formatted_row = max(self.max_formatted_row, max_row)

    def _set_basic_info(self):
        self.ws.title = f"Screening(FY{self.start_year - 2000}{self.end_year - 2000})"
        self.ws['A1'] = f"분석대상법인: {self.tested_party}"
        self.ws['A2'] = f"분석대상연도: FY{self.start_year}-{self.end_year}"
        self.ws['A3'] = f"{self.name}"

    def _set_quantitative_criteria_table(self):
        self.ws['A5'] = "양적기준"
        for i in range(self.number_of_criteria):
            self.ws.cell(row=6 + i, column=1).value = i + 1

        q_cond_row = self.quantitative_start_row
        q_cond_col = self.quantitative_start_col
        self.ws.cell(row=q_cond_row, column=q_cond_col).value = "양적조건"
        self.ws.merge_cells(start_row=q_cond_row, end_row=q_cond_row,
                            start_column=q_cond_col, end_column=q_cond_col + self.number_of_criteria - 1)
        self.ws.cell(row=q_cond_row, column=q_cond_col).fill = self.color_code["green"]

        for i in range(self.number_of_criteria):
            target_col = q_cond_col + i
            self.ws.cell(row=q_cond_row + 1, column=target_col).value = f"기준{i + 1}"
            
            if i < len(self.criteria_list):
                 account_name = self.criteria_list[i].get('account', '')
                 self.ws.cell(row=q_cond_row + 2, column=target_col).value = account_name

            self.ws.cell(row=q_cond_row + 1, column=target_col).fill = self.color_code["green"]

        q_pass_col = q_cond_col + self.number_of_criteria
        self.ws.cell(row=q_cond_row + 1, column=q_pass_col).value = "양적통과"
        self.ws.merge_cells(start_row=q_cond_row + 1, end_row=q_cond_row + 2,
                            start_column=q_pass_col, end_column=q_pass_col)
        
        for r_idx in range(q_cond_row, q_cond_row + 3):
            for c_idx in range(q_cond_col, q_pass_col + 1):
                cell = self.ws.cell(row=r_idx, column=c_idx)
                if not cell.fill.start_color.rgb == self.color_code["green"].start_color.rgb:
                     cell.fill = self.color_code["green"]
    
        for i in range(self.number_of_criteria):
            self.ws.cell(row=q_cond_row + 2, column=q_cond_col + i).fill = self.color_code["green"]

    def _set_qualitative_criteria_table(self):
        q_data = self._get_qualitative_criteria_data()
        q_keys = self._get_qualitative_criteria_keys()
        
        q_start_row = self.qualitative_start_row
        q_start_col = self.qualitative_start_col

        self.ws.cell(row=q_start_row, column=q_start_col).value = "질적조건"
        self.ws.merge_cells(start_row=q_start_row, end_row=q_start_row,
                            start_column=q_start_col, end_column=q_start_col + len(q_keys) - 1)
        self.ws.cell(row=q_start_row, column=q_start_col).fill = self.color_code['yellow']

        for col_idx, key in enumerate(q_keys):
            target_col = q_start_col + col_idx
            self.ws.cell(row=q_start_row + 1, column=target_col).value = key
            self.ws.cell(row=q_start_row + 2, column=target_col).value = q_data[key]

            if q_data[key] == "":
                self.ws.merge_cells(start_row=q_start_row + 1, end_row=q_start_row + 2,
                                    start_column=target_col, end_column=target_col)
            
            self.ws.cell(row=q_start_row + 1, column=target_col).fill = self.color_code['yellow']
            self.ws.cell(row=q_start_row + 2, column=target_col).fill = self.color_code['yellow']

    def _set_selection_history_columns(self):
        s_hist_list = ["FY2023 BEPS 선정이력", "최종선정/제외 Comment", "선정여부 추가검토"]
        s_hist_row = self.qualitative_start_row
        s_hist_col = self.selection_history_start_col

        for col_idx, item in enumerate(s_hist_list):
            target_col = s_hist_col + col_idx
            self.ws.cell(row=s_hist_row, column=target_col).value = item
            self.ws.cell(row=s_hist_row, column=target_col).fill = self.color_code["orange"]
            self.ws.merge_cells(start_row=s_hist_row, end_row=s_hist_row + 2,
                            start_column=target_col, end_column=target_col)

    def _set_unadjusted_columns(self):
        unadj_row = self.qualitative_start_row
        unadj_col = self.unadjusted_start_col
        unadj_list = self._get_unadj_list()
        
        num_cols_per_metric = len(unadj_list)
        total_unadj_cols = num_cols_per_metric * 3

        self.ws.cell(row=unadj_row, column=unadj_col).value = "Unadjusted"
        self.ws.merge_cells(start_row=unadj_row, end_row=unadj_row,
                            start_column=unadj_col, end_column=unadj_col + total_unadj_cols - 1)
        self.ws.cell(row=unadj_row, column=unadj_col).fill = self.color_code['green']

        metrics = ['OM', 'MTC', 'BR']
        for metric_idx, metric_name in enumerate(metrics):
            metric_start_col = unadj_col + metric_idx * num_cols_per_metric
            self.ws.cell(row=unadj_row + 1, column=metric_start_col).value = metric_name
            self.ws.merge_cells(start_row=unadj_row + 1, end_row=unadj_row + 1,
                                start_column=metric_start_col, 
                                end_column=metric_start_col + num_cols_per_metric - 1)
            self.ws.cell(row=unadj_row + 1, column=metric_start_col).fill = self.color_code['green']

            for col_idx, item in enumerate(unadj_list):
                target_col = metric_start_col + col_idx
                self.ws.cell(row=unadj_row + 2, column=target_col).value = item
                self.ws.cell(row=unadj_row + 2, column=target_col).fill = self.color_code['green']

    def _set_wa3_columns(self):
        wa3_row = self.qualitative_start_row
        wa3_col = self.wa3_start_col
        wa3_list = self._get_wa3_list()
        ratio_tab_list = self._get_ratio_tab_list()

        self.ws.cell(row=wa3_row, column=wa3_col).value = f"FY{self.start_year-2000}-FY{self.end_year-2000} {self.num_years}개년 평균"
        self.ws.merge_cells(start_row=wa3_row, end_row=wa3_row + 1,
                            start_column=wa3_col, end_column=wa3_col + len(wa3_list) - 1)
        self.ws.cell(row=wa3_row, column=wa3_col).fill = self.color_code['orange']

        for col_idx, item in enumerate(wa3_list):
            target_col = wa3_col + col_idx
            self.ws.cell(row=wa3_row + 2, column=target_col).value = item
            self.ws.cell(row=wa3_row + 2, column=target_col).fill = self.color_code['orange']

        ratio_col_start = wa3_col + len(wa3_list)
        for col_idx, item in enumerate(ratio_tab_list):
            target_col = ratio_col_start + col_idx
            self.ws.cell(row=wa3_row, column=target_col).value = item
            self.ws.cell(row=wa3_row, column=target_col).fill = self.color_code['orange']
            self.ws.merge_cells(start_row=wa3_row, end_row=wa3_row + 2,
                                start_column=target_col, end_column=target_col)

    def _set_raw_data_columns(self):
        raw_data_row = self.qualitative_start_row
        raw_data_col = self.raw_data_start_col

        self.ws.cell(row=raw_data_row, column=1).value = "#"
        self.ws.merge_cells(start_row=raw_data_row, end_row=raw_data_row + 2, start_column=1, end_column=1)
        self.ws.cell(row=raw_data_row, column=1).fill = self.color_code['orange']

        self.ws.cell(row=raw_data_row, column=2).value = "BvD ID number"
        self.ws.cell(row=raw_data_row + 2, column=2).value = "BvD ID number"
        self.ws.merge_cells(start_row=raw_data_row, end_row=raw_data_row + 1, start_column=2, end_column=2)
        self.ws.cell(row=raw_data_row, column=2).fill = self.color_code['orange']
        self.ws.cell(row=raw_data_row+2, column=2).fill = self.color_code['orange']

        self.ws.cell(row=raw_data_row, column=3).value = "Company name Latin alphabet"
        self.ws.cell(row=raw_data_row + 2, column=3).value = "Company name"
        self.ws.merge_cells(start_row=raw_data_row, end_row=raw_data_row + 1, start_column=3, end_column=3)
        self.ws.cell(row=raw_data_row, column=3).fill = self.color_code['orange']
        self.ws.cell(row=raw_data_row+2, column=3).fill = self.color_code['orange']

        for col_idx, item in enumerate(self.ordered_columns):
            target_col = raw_data_col + col_idx
            self.ws.cell(row=raw_data_row, column=target_col).value = item
            self.ws.cell(row=raw_data_row, column=target_col).fill = self.color_code["gray"]
            self.ws.merge_cells(start_row=raw_data_row, end_row=raw_data_row + 2,
                                start_column=target_col, end_column=target_col)
                                
    def _set_flow_columns(self):
        flow_row = self.qualitative_start_row
        flow_col = self.flow_start_col
        flow_list = ["매출채권 (Flow)", "매입채무 (Flow)", "재고자산 (Flow)", "무형자산 (Flow)", "유형자산 (Flow)", "총자산 (Flow)"]

        num_flow_cols = self.num_years + 1
        total_flow_cols = len(flow_list) * num_flow_cols

        self.ws.cell(row=flow_row, column=flow_col).value = "FLOW"
        self.ws.merge_cells(start_row=flow_row, end_row=flow_row,
                            start_column=flow_col, end_column=flow_col + total_flow_cols - 1)
        self.ws.cell(row=flow_row, column=flow_col).fill = self.color_code['orange']

        for col_idx, item in enumerate(flow_list):
            start_current_flow_col = flow_col + col_idx * num_flow_cols
            self.ws.cell(row=flow_row + 1, column=start_current_flow_col).value = item
            self.ws.merge_cells(start_row=flow_row + 1, end_row=flow_row + 1,
                                start_column=start_current_flow_col, 
                                end_column=start_current_flow_col + num_flow_cols - 1)
            self.ws.cell(row=flow_row + 1, column=start_current_flow_col).fill = self.color_code['orange']

            for i in range(num_flow_cols):
                target_cell = self.ws.cell(row=flow_row + 2, column=start_current_flow_col + i)
                if i == num_flow_cols - 1:
                    target_cell.value = f"{self.num_years}WA"
                else:
                    target_cell.value = f"FY{self.start_year - 2000 + i}" 
                target_cell.fill = self.color_code['orange']
        
        self.max_formatted_col = max(self.max_formatted_col, flow_col + total_flow_cols - 1)
        self.max_formatted_row = max(self.max_formatted_row, flow_row + 2)

    def _populate_raw_data_from_excel(self):
        if not self.data_path:
            print("Error: data_path가 설정되지 않았습니다. Excel 파일 경로를 지정해주세요.")
            return

        try:
            # Results 시트만 읽기
            source_df = pd.read_excel(
                self.data_path,
                sheet_name="Results",
                header=0
            )
        except FileNotFoundError:
            print(f"Error: 파일 '{self.data_path}'을(를) 찾을 수 없습니다.")
            return
        except Exception as e:
            print(f"Error: Excel 파일을 읽는 중 오류 발생: {e}")
            return

        # =========================
        # 2번째 행 제거 (무가치한 헤더)
        # =========================
        if len(source_df) >= 1:
            source_df = source_df.drop(index=0).reset_index(drop=True)

        # =========================
        # 데이터 시작 행 계산
        # =========================
        data_body_start_row = self.qualitative_start_row + 3

        # =========================
        # Row number 채우기
        # =========================
        for i in range(len(source_df)):
            self.ws.cell(row=data_body_start_row + i, column=1).value = i + 1

        # =========================
        # 컬럼별 데이터 매핑
        # =========================
        for col_idx, target_col_name in enumerate(self.ordered_columns):
            sheet_col_num = self.raw_data_start_col + col_idx

            if target_col_name in source_df.columns:
                source_series = source_df[target_col_name]

                for row_idx, value in enumerate(source_series):
                    if pd.isna(value):
                        value = None

                    self.ws.cell(
                        row=data_body_start_row + row_idx,
                        column=sheet_col_num
                    ).value = value
            else:
                print(
                    f"Warning: 원본 Excel 파일 '{self.data_path}'의 Results 시트에 "
                    f"'{target_col_name}' 컬럼이 없습니다."
                )

    # def _populate_raw_data_from_excel(self):
    #     if not self.data_path:
    #         print("Error: data_path가 설정되지 않았습니다. Excel 파일 경로를 지정해주세요.")
    #         return
        
    #     try:
    #         source_df = pd.read_excel(self.data_path)
    #     except FileNotFoundError:
    #         print(f"Error: 파일 '{self.data_path}'을(를) 찾을 수 없습니다.")
    #         return
    #     except Exception as e:
    #         print(f"Error: Excel 파일을 읽는 중 오류 발생: {e}")
    #         return

    #     data_body_start_row = self.qualitative_start_row + 3

    #     for i in range(len(source_df)):
    #         self.ws.cell(row=data_body_start_row + i, column=1).value = i + 1

    #     for col_idx, target_col_name in enumerate(self.ordered_columns):
    #         sheet_col_num = self.raw_data_start_col + col_idx 
            
    #         if target_col_name in source_df.columns:
    #             source_series = source_df[target_col_name]
    #             for row_idx, value in enumerate(source_series):
    #                 if pd.isna(value):
    #                     value = None
    #                 cell_to_write = self.ws.cell(row=data_body_start_row + row_idx, column=sheet_col_num)
    #                 cell_to_write.value = value
    #         else:
    #             print(f"Warning: 원본 Excel 파일 '{self.data_path}'에 '{target_col_name}' 컬럼이 없습니다.")

    def create_format(self):
        self._set_basic_info()
        self._set_quantitative_criteria_table()
        self._set_qualitative_criteria_table()
        self._set_selection_history_columns()
        self._set_unadjusted_columns()
        self._set_wa3_columns()
        self._set_raw_data_columns()
        self._set_flow_columns() 

        self._apply_common_styles(min_row=self.qualitative_start_row,
                                max_row=self.qualitative_start_row + 2,
                                min_col=1,
                                max_col=self.max_formatted_col)

    def insert_formular(self):
        """수식을 동적으로 생성하여 삽입합니다."""
        asset_list = [
            "Debtors\nth USD ",
            "Creditors\nth USD ",
            "Stock\nth USD ",
            "Intangible assets\nth USD ",
            "Tangible fixed assets\nth USD ",
            "Total assets\nth USD "
        ]
        
        col_name = {}
        for asset in asset_list:
            for year in range(self.start_year-1, self.end_year+1):
                col_name[asset+str(year)] = self.raw_col_alphabet[asset+str(year)]
        
        num_flow_cols = self.num_years + 1
        
        for asset in asset_list:
            asset_idx = asset_list.index(asset)
            
            for year_idx in range(self.num_years):
                col = self.flow_start_col + asset_idx * num_flow_cols + year_idx
                for row in range(self.qualitative_start_row + 3, self.ws.max_row + 1):
                    cell = self.ws.cell(row=row, column=col)
                    prev_year = self.start_year - 1 + year_idx
                    curr_year = self.start_year + year_idx
                    cell.value = f"=IFERROR(SUM({col_name[asset+str(prev_year)]}{row}:{col_name[asset+str(curr_year)]}{row})/2,0)"
            
            wa_col = self.flow_start_col + asset_idx * num_flow_cols + self.num_years
            for row in range(self.qualitative_start_row + 3, self.ws.max_row + 1):
                cell = self.ws.cell(row=row, column=wa_col)
                start_col_letter = get_column_letter(wa_col - self.num_years)
                end_col_letter = get_column_letter(wa_col - 1)
                cell.value = f"=IFERROR(SUM({start_col_letter}{row}:{end_col_letter}{row})/{self.num_years},0)"
                
                if "Stock" in asset:
                    self.ws.cell(row=row, column=self.wa3_start_col+3).value = f"=IFERROR({get_column_letter(wa_col)}{row},0)"
                elif "Intangible" in asset:
                    self.ws.cell(row=row, column=self.wa3_start_col+5).value = f"=IFERROR({get_column_letter(wa_col)}{row},0)"
                elif "Tangible" in asset:
                    self.ws.cell(row=row, column=self.wa3_start_col+6).value = f"=IFERROR({get_column_letter(wa_col)}{row},0)"
                elif "Total" in asset:
                    self.ws.cell(row=row, column=self.wa3_start_col+7).value = f"=IFERROR({get_column_letter(wa_col)}{row},0)"
        
        pl_list = {
            f"Operating revenue (Turnover)\nth USD {self.start_year}": 0,
            f"Operating profit (loss) [EBIT]\nth USD {self.start_year}": 1,
            f"Other operating expense (income)\nth USD {self.start_year}": 2,
            f"Research & Development expenses\nth USD {self.start_year}": 4,
            f"Costs of goods sold\nth USD {self.start_year}": 8,
            f"Number of employees\n{self.start_year}": 9,
        }

        for pl, col_idx in pl_list.items():
            for row in range(self.qualitative_start_row + 3, self.ws.max_row + 1):
                cell = self.ws.cell(row=row, column=self.wa3_start_col + col_idx)
                start_col = self.raw_col_number[pl]
                end_col = start_col + self.num_years - 1
                cell.value = f"=IFERROR(SUM({get_column_letter(start_col)}{row}:{get_column_letter(end_col)}{row})/{self.num_years},0)"
        
        ratio_idx = {1:(4,0), 2:(2,0), 3:(5,7), 4:(6,7), 5:(3,7), 6:(3,8)}

        for col_idx, (numerator, denominator) in ratio_idx.items():
            for row in range(self.qualitative_start_row + 3, self.ws.max_row + 1):
                cell = self.ws.cell(row=row, column=self.wa3_start_col + 9 + col_idx)
                if col_idx == 6:
                    cell.value = f'=IFERROR(365/({get_column_letter(self.wa3_start_col+denominator)}{row}/{get_column_letter(self.wa3_start_col+numerator)}{row}), "")'
                else:
                    cell.value = f"=IFERROR({get_column_letter(self.wa3_start_col+numerator)}{row}/{get_column_letter(self.wa3_start_col+denominator)}{row},0)"

        num_cols_per_metric = len(self._get_unadj_list())
        
        for year_idx in range(self.num_years):
            col = self.unadjusted_start_col + year_idx
            for row in range(self.qualitative_start_row + 3, self.ws.max_row + 1):
                cell = self.ws.cell(row=row, column=col)
                year = self.start_year + year_idx
                op_col = self.raw_col_alphabet[f"Operating profit (loss) [EBIT]\nth USD {year}"]
                rev_col = self.raw_col_alphabet[f"Operating revenue (Turnover)\nth USD {year}"]
                cell.value = f"=IFERROR({op_col}{row}/{rev_col}{row},0)"
        
        avg_col = self.unadjusted_start_col + self.num_years
        for row in range(self.qualitative_start_row + 3, self.ws.max_row + 1):
            cell = self.ws.cell(row=row, column=avg_col)
            op_start = self.raw_col_alphabet[f"Operating profit (loss) [EBIT]\nth USD {self.start_year}"]
            op_end = self.raw_col_alphabet[f"Operating profit (loss) [EBIT]\nth USD {self.end_year}"]
            rev_start = self.raw_col_alphabet[f"Operating revenue (Turnover)\nth USD {self.start_year}"]
            rev_end = self.raw_col_alphabet[f"Operating revenue (Turnover)\nth USD {self.end_year}"]
            cell.value = f"=IFERROR(SUM({op_start}{row}:{op_end}{row})/SUM({rev_start}{row}:{rev_end}{row}),0)"
        
        maxmin_col = self.unadjusted_start_col + self.num_years + 1
        
        # 1. BvD ID number & Company name reference
        # Raw Data의 BvD ID, Name 컬럼 인덱스 (raw_data_start_col, raw_data_start_col+1)
        raw_bvd_col = get_column_letter(self.raw_data_start_col)
        raw_name_col = get_column_letter(self.raw_data_start_col + 1)
        
        for row in range(self.qualitative_start_row + 3, self.ws.max_row + 1):
            # BvD ID (Col 2)
            self.ws.cell(row=row, column=2).value = f"={raw_bvd_col}{row}"
            # Company Name (Col 3)
            self.ws.cell(row=row, column=3).value = f"={raw_name_col}{row}"

        for row in range(self.qualitative_start_row + 3, self.ws.max_row + 1):
            cell = self.ws.cell(row=row, column=maxmin_col)
            start_letter = get_column_letter(self.unadjusted_start_col)
            end_letter = get_column_letter(self.unadjusted_start_col + self.num_years - 1)
            cell.value = f"=IFERROR(MAX({start_letter}{row}:{end_letter}{row})-MIN({start_letter}{row}:{end_letter}{row}),0)"
        
        mtc_start = self.unadjusted_start_col + num_cols_per_metric
        for year_idx in range(self.num_years):
            col = mtc_start + year_idx
            for row in range(self.qualitative_start_row + 3, self.ws.max_row + 1):
                cell = self.ws.cell(row=row, column=col)
                year = self.start_year + year_idx
                op_col = self.raw_col_alphabet[f"Operating profit (loss) [EBIT]\nth USD {year}"]
                rev_col = self.raw_col_alphabet[f"Operating revenue (Turnover)\nth USD {year}"]
                cell.value = f"=IFERROR({op_col}{row}/({rev_col}{row}-{op_col}{row}),0)"
        
        avg_col = mtc_start + self.num_years
        for row in range(self.qualitative_start_row + 3, self.ws.max_row + 1):
            cell = self.ws.cell(row=row, column=avg_col)
            op_start = self.raw_col_alphabet[f"Operating profit (loss) [EBIT]\nth USD {self.start_year}"]
            op_end = self.raw_col_alphabet[f"Operating profit (loss) [EBIT]\nth USD {self.end_year}"]
            rev_start = self.raw_col_alphabet[f"Operating revenue (Turnover)\nth USD {self.start_year}"]
            rev_end = self.raw_col_alphabet[f"Operating revenue (Turnover)\nth USD {self.end_year}"]
            cell.value = f"=IFERROR(SUM({op_start}{row}:{op_end}{row})/(SUM({rev_start}{row}:{rev_end}{row})-SUM({op_start}{row}:{op_end}{row})),0)"
        
        maxmin_col = mtc_start + self.num_years + 1
        for row in range(self.qualitative_start_row + 3, self.ws.max_row + 1):
            cell = self.ws.cell(row=row, column=maxmin_col)
            start_letter = get_column_letter(mtc_start)
            end_letter = get_column_letter(mtc_start + self.num_years - 1)
            cell.value = f"=IFERROR(MAX({start_letter}{row}:{end_letter}{row})-MIN({start_letter}{row}:{end_letter}{row}),0)"
        
        br_start = self.unadjusted_start_col + num_cols_per_metric * 2
        for year_idx in range(self.num_years):
            col = br_start + year_idx
            for row in range(self.qualitative_start_row + 3, self.ws.max_row + 1):
                cell = self.ws.cell(row=row, column=col)
                year = self.start_year + year_idx
                gp_col = self.raw_col_alphabet[f"Gross profit\nth USD {year}"]
                opex_col = self.raw_col_alphabet[f"Other operating expense (income)\nth USD {year}"]
                cell.value = f"=IFERROR({gp_col}{row}/{opex_col}{row},0)"
        
        avg_col = br_start + self.num_years
        for row in range(self.qualitative_start_row + 3, self.ws.max_row + 1):
            cell = self.ws.cell(row=row, column=avg_col)
            gp_start = self.raw_col_alphabet[f"Gross profit\nth USD {self.start_year}"]
            gp_end = self.raw_col_alphabet[f"Gross profit\nth USD {self.end_year}"]
            opex_start = self.raw_col_alphabet[f"Other operating expense (income)\nth USD {self.start_year}"]
            opex_end = self.raw_col_alphabet[f"Other operating expense (income)\nth USD {self.end_year}"]
            cell.value = f"=IFERROR(SUM({gp_start}{row}:{gp_end}{row})/SUM({opex_start}{row}:{opex_end}{row}),0)"
        
        maxmin_col = br_start + self.num_years + 1
        for row in range(self.qualitative_start_row + 3, self.ws.max_row + 1):
            cell = self.ws.cell(row=row, column=maxmin_col)
            start_letter = get_column_letter(br_start)
            end_letter = get_column_letter(br_start + self.num_years - 1)
            cell.value = f"=IFERROR(MAX({start_letter}{row}:{end_letter}{row})-MIN({start_letter}{row}:{end_letter}{row}),0)"

    def insert_pass_fail_summary(self):
        """
        C20, C21에 탈락/통과 텍스트 입력 및 수식 적용
        각 기준별 누적 통과/탈락 통계를 계산
        """
        # 1. 텍스트 입력
        self.ws['C20'] = "탈락"
        self.ws['C21'] = "통과"
        self.ws['C20'].border = THIN_BORDER
        self.ws['C21'].border = THIN_BORDER
        
        # 2. 수식 입력
        data_start_row = self.qualitative_start_row + 3
        # 데이터가 없을 경우를 대비해 max_row 체크
        if self.ws.max_row < data_start_row:
             data_end_row = data_start_row + 10 # 임의의 범위
        else:
             data_end_row = self.ws.max_row

        start_col = self.quantitative_start_col
        
        for i in range(self.number_of_criteria):
            target_col_idx = start_col + i
            target_col_letter = get_column_letter(target_col_idx)
            
            # 범위 문자열 (예: D25:D100)
            range_str = f"${target_col_letter}${data_start_row}:${target_col_letter}${data_end_row}"
            
            # 누적으로 이전 기준들의 "Yes" 조건을 모두 포함해야 함
            # COUNTIFS(Criteria1, "Yes", Criteria2, "Yes", ..., TargetCriteria, "No/Yes")
            
            conditions = []
            for j in range(i + 1):
                prev_col_letter = get_column_letter(start_col + j)
                prev_range = f"${prev_col_letter}${data_start_row}:${prev_col_letter}${data_end_row}"
                conditions.append(prev_range)
                conditions.append('"Yes"') # 일단 모두 Yes로 추가해두고 마지막만 수정
            
            # --- 탈락 수식 (Row 20) ---
            # 마지막 조건은 "No"여야 함
            fail_conditions = list(conditions)
            fail_conditions[-1] = '"No"'
            
            fail_formula = f'=COUNTIFS({",".join(fail_conditions)})'
            self.ws.cell(row=20, column=target_col_idx).value = fail_formula
            self.ws.cell(row=20, column=target_col_idx).border = THIN_BORDER
            
            # --- 통과 수식 (Row 21) ---
            # 마지막 조건도 "Yes"여야 함 (이미 conditions가 모두 "Yes"임)
            pass_formula = f'=COUNTIFS({",".join(conditions)})'
            self.ws.cell(row=21, column=target_col_idx).value = pass_formula
            self.ws.cell(row=21, column=target_col_idx).border = THIN_BORDER
            

    def apply_quantitative_criteria_formulas(self, criteria_configs):
        """
        양적기준 수식을 실제 셀에 적용합니다.
        
        Parameters:
        - criteria_configs: 기준 설정 리스트
          각 항목은 딕셔너리 형태:
          {
              'type': 'text' | 'numeric' | 'ratio' | 'data_availability',
              'field_name': 필드명,
              'condition_type': 조건 타입,
              'value': 비교값 (텍스트) 또는 threshold (숫자),
              'include': True | False,
              'use_threshold_cell': True | False (선택),
              'count_requirement': None | 'all' | 'any' | int (선택)
          }
        """
        if not criteria_configs:
            print("Warning: 양적기준 설정이 비어있습니다.")
            return
        
        if len(criteria_configs) > self.number_of_criteria:
            print(f"Warning: 설정된 기준({len(criteria_configs)})이 number_of_criteria({self.number_of_criteria})보다 많습니다.")
            criteria_configs = criteria_configs[:self.number_of_criteria]
        
        # 데이터 시작 행 (헤더 3줄 아래)
        data_start_row = self.qualitative_start_row + 3
        
        # Raw 데이터가 있는 경우에만 수식 적용
        if self.ws.max_row < data_start_row:
            print("Warning: Raw 데이터가 없습니다. 먼저 _populate_raw_data_from_excel()을 실행하세요.")
            return
        
        # 각 기준에 대해 수식 생성 및 적용
        for criteria_idx, config in enumerate(criteria_configs):
            criteria_col = self.quantitative_start_col + criteria_idx
            
            # 각 데이터 행에 수식 적용
            for row in range(data_start_row, self.ws.max_row + 1):
                formula = self._generate_formula_from_config(config, row, criteria_idx + 1)
                cell = self.ws.cell(row=row, column=criteria_col)
                cell.value = formula
        
        # 양적통과 컬럼 (모든 기준을 통과한 경우만 Yes)
        self._apply_quantitative_pass_formula(data_start_row)
        
        print(f"양적기준 수식 적용 완료: {len(criteria_configs)}개 기준, {self.ws.max_row - data_start_row + 1}개 행")
    
    def _generate_formula_from_config(self, config, row_number, criteria_index):
        """설정 딕셔너리로부터 수식을 생성합니다."""
        criteria_type = config.get('type')
        
        if criteria_type == 'text':
            return self.formula_generator.generate_text_criteria(
                field_name=config.get('field_name'),
                condition_type=config.get('condition_type'),
                value=config.get('value', ''),
                row_number=row_number,
                include=config.get('include', True)
            )
        
        elif criteria_type == 'numeric':
            return self.formula_generator.generate_numeric_criteria(
                field_name=config.get('field_name'),
                condition_type=config.get('condition_type'),
                threshold=config.get('value', 0),
                row_number=row_number,
                include=config.get('include', True),
                use_threshold_cell=config.get('use_threshold_cell', False),
                criteria_index=criteria_index if config.get('use_threshold_cell', False) else None,
                count_requirement=config.get('count_requirement', None)
            )
        
        elif criteria_type == 'ratio':
            return self.formula_generator.generate_ratio_criteria(
                ratio_name=config.get('field_name'),
                condition_type=config.get('condition_type'),
                threshold=config.get('value', 0),
                row_number=row_number,
                include=config.get('include', True),
                use_threshold_cell=config.get('use_threshold_cell', False),
                criteria_index=criteria_index if config.get('use_threshold_cell', False) else None
            )
        
        elif criteria_type == 'data_availability':
            return self.formula_generator.generate_data_availability_criteria(
                field_names=config.get('field_names', []),
                row_number=row_number,
                include=config.get('include', True)
            )
        
        elif criteria_type == 'wa3':
            return self.formula_generator.generate_wa3_numeric_criteria(
                ratio_name=config.get('field_name'),
                condition_type=config.get('condition_type'),
                threshold=config.get('value', 0),
                row_number=row_number,
                include=config.get('include', True),
                use_threshold_cell=config.get('use_threshold_cell', False),
                criteria_index=criteria_index if config.get('use_threshold_cell', False) else None
            )
        
        else:
            return '="Error: Unknown criteria type"'
    
    def _apply_quantitative_pass_formula(self, data_start_row):
        """양적통과 컬럼에 수식을 적용합니다 (모든 기준이 Yes인 경우만 Yes)."""
        pass_col = self.quantitative_start_col + self.number_of_criteria
        
        for row in range(data_start_row, self.ws.max_row + 1):
            # 모든 양적기준 셀 참조
            criteria_cells = []
            for i in range(self.number_of_criteria):
                col_letter = get_column_letter(self.quantitative_start_col + i)
                criteria_cells.append(f'{col_letter}{row}')
            
            # COUNTIF를 사용하여 모든 기준이 "Yes"인지 확인
            criteria_range_start = criteria_cells[0].replace(str(row), '')
            criteria_range_end = criteria_cells[-1].replace(str(row), '')
            formula = f'=IF(COUNTIF(${criteria_range_start}${row}:${criteria_range_end}${row},"Yes")={self.number_of_criteria},"Yes","No")'
            
            cell = self.ws.cell(row=row, column=pass_col)
            cell.value = formula

    def apply_final_styles(self):
        """
        최종적으로 서식을 적용합니다.
        1. Unadjusted, WA3(Ratio): 0.00%
        2. WA3(Metrics), Raw Data(Turnover~), Flow: Accounting Format
        3. All Borders
        """
        ACCOUNTING_FORMAT = '_(* #,##0_);_(* (#,##0);_(* "-"??_);_(@_)'
        PERCENTAGE_FORMAT = '0.00%'
        
        # 데이터 영역 (헤더 제외)
        data_start_row = self.qualitative_start_row + 3
        max_row = self.ws.max_row
        max_col = self.max_formatted_col
        
        # 1. Unadjusted (Percentage)
        for col in range(self.unadjusted_start_col, self.unadjusted_start_col + self.unadjusted_num_cols):
            for row in range(data_start_row, max_row + 1):
                self.ws.cell(row=row, column=col).number_format = PERCENTAGE_FORMAT
                
        # 2. WA3 Metrics (Accounting)
        # 종업원수 포함하려면 wa3_list 전체 적용
        wa3_metrics_count = len(self._get_wa3_list())
        for col in range(self.wa3_start_col, self.wa3_start_col + wa3_metrics_count):
            for row in range(data_start_row, max_row + 1):
                self.ws.cell(row=row, column=col).number_format = ACCOUNTING_FORMAT

        # 3. WA3 Ratios (Percentage)
        wa3_ratios_count = len(self._get_ratio_tab_list())
        ratio_start_col = self.wa3_start_col + wa3_metrics_count
        for col in range(ratio_start_col, ratio_start_col + wa3_ratios_count):
            for row in range(data_start_row, max_row + 1):
                self.ws.cell(row=row, column=col).number_format = PERCENTAGE_FORMAT
                
        # 4. Raw Data (Accounting from Turnover)
        # Turnover 컬럼 인덱스 찾기
        turnover_col_name = f"Operating revenue (Turnover)\nth USD "
        # ordered_columns에서 Turnover의 인덱스 확인
        try:
            # Turnover가 포함된 첫 컬럼 인덱스를 찾음 (년도별 컬럼 중 첫번째)
            # 여기서는 ordered_columns에 'Operating revenue (Turnover)\nth USD 2021' 처럼 년도가 붙어있으므로
            # 기본 prefix로 검색
            
            # Turnover가 시작되는 첫 컬럼 찾기
            turnover_start_index = -1
            for idx, col_name in enumerate(self.ordered_columns):
                if "Operating revenue (Turnover)" in col_name:
                    turnover_start_index = idx
                    break
            
            if turnover_start_index != -1:
                abs_turnover_col = self.raw_data_start_col + turnover_start_index
                # Raw Data 끝까지 적용
                raw_data_end_col = self.raw_data_start_col + len(self.ordered_columns)
                
                for col in range(abs_turnover_col, raw_data_end_col):
                    for row in range(data_start_row, max_row + 1):
                        self.ws.cell(row=row, column=col).number_format = ACCOUNTING_FORMAT
        except Exception as e:
            print(f"Warning: Raw Data 서식 적용 중 오류 발생: {e}")

        # 5. Flow (Accounting)
        # Flow 시작부터 max_formatted_col까지 (Flow가 마지막 부분이라 가정)
        # 정확히는 flow_start_col 부터 flow 끝까지
        flow_list = ["매출채권 (Flow)", "매입채무 (Flow)", "재고자산 (Flow)", "무형자산 (Flow)", "유형자산 (Flow)", "총자산 (Flow)"]
        num_flow_cols = self.num_years + 1
        total_flow_cols = len(flow_list) * num_flow_cols
        
        for col in range(self.flow_start_col, self.flow_start_col + total_flow_cols):
             for row in range(data_start_row, max_row + 1):
                self.ws.cell(row=row, column=col).number_format = ACCOUNTING_FORMAT
        
        # 6. All Borders
        # 전체 테이블 범위: qualitative_start_row 부터 max_row, 1부터 max_col
        # max_formatted_col이 정확하지 않을 수 있으므로 max_col 재계산
        final_max_col = max(self.max_formatted_col, self.flow_start_col + total_flow_cols - 1)
        
        for row in range(self.qualitative_start_row, max_row + 1):
            for col in range(1, final_max_col + 1):
                cell = self.ws.cell(row=row, column=col)
                cell.border = THIN_BORDER

    def save_file(self):
        self.wb.save(f"{self.name}.xlsx")



import pandas as pd
from openpyxl.utils import get_column_letter
class SimpleUserInputConverter:
   """
   단순화된 사용자 입력 변환 클래스
   Account만 받아서 자동으로 Type과 계산구분 결정
   """
   # Account별 자동 설정
   ACCOUNT_CONFIG = {
       # 텍스트 타입
       "감사의견": {
           'type': 'text',
           'field_name': 'Audit status\n',
           'condition_type': 'all_equals',
           'default_value': 'Unqualified',
           'default_include': True
       },
       "상장여부": {
           'type': 'text',
           'field_name': 'Listing status',
           'condition_type': 'equals',
           'default_value': 'Listed',
           'default_include': True
       },
       # 가용성
       "재무정보가용성": {
           'type': 'data_availability',
           'field_names': [
               "Operating revenue (Turnover)\nth USD ",
               "Gross profit\nth USD ",
               "Operating profit (loss) [EBIT]\nth USD "
           ],
           'default_include': True
       },
       # 평균값 (WA3) - 금액
       "매출액(평균)": {
           'type': 'wa3',
           'field_name': '매출액',
           'condition_type': 'gt',
           'default_value': 0,
           'default_include': True
       },
       "영업이익(평균)": {
           'type': 'wa3',
           'field_name': '영업이익',
           'condition_type': 'lt',
           'default_value': 0,
           'default_include': False
       },
       "판매및관리비(평균)": {  # 영업비용
           'type': 'wa3',
           'field_name': '영업비용',
           'condition_type': 'gt',
           'default_value': 0,
           'default_include': False
       },
       "재고자산(평균)": {
           'type': 'wa3',
           'field_name': '재고자산',
           'condition_type': 'gt',
           'default_value': 0,
           'default_include': False
       },
       "연구개발비(평균)": {
           'type': 'wa3',
           'field_name': '연구개발비',
           'condition_type': 'gt',
           'default_value': 0,
           'default_include': False
       },
       "무형자산(평균)": {
           'type': 'wa3',
           'field_name': '무형자산',
           'condition_type': 'gt',
           'default_value': 0,
           'default_include': False
       },
       "유형자산(평균)": {
           'type': 'wa3',
           'field_name': '유형자산',
           'condition_type': 'gt',
           'default_value': 0,
           'default_include': False
       },
       "총자산(평균)": {
           'type': 'wa3',
           'field_name': '총자산',
           'condition_type': 'gt',
           'default_value': 0,
           'default_include': False
       },
       "매출원가(평균)": {
           'type': 'wa3',
           'field_name': '매출원가',
           'condition_type': 'gt',
           'default_value': 0,
           'default_include': False
       },
       # 개별값 - 금액 (1개년이라도)
       "영업이익(1개년이라도)": {
           'type': 'numeric',
           'field_name': 'Operating profit (loss) [EBIT]\nth USD ',
           'condition_type': 'lt',
           'count_requirement': 'any',
           'default_value': 0,
           'default_include': False
       },
       # 개별값 - 금액 (3년연속 = 모든 연도)
       "영업이익(연속)": {
           'type': 'numeric',
           'field_name': 'Operating profit (loss) [EBIT]\nth USD ',
           'condition_type': 'lt',
           'count_requirement': 'all',
           'default_value': 0,
           'default_include': False
       },
       # 비율 (평균)
       "연구개발비/매출액(평균)": {
           'type': 'ratio',
           'field_name': '연구개발비/매출액',
           'condition_type': 'gt',
           'default_value': 0,
           'default_include': False
       },
       "영업비용/매출액(평균)": {
           'type': 'ratio',
           'field_name': '영업비용/매출액',
           'condition_type': 'gt',
           'default_value': 0,
           'default_include': False
       },
       "무형자산/총자산(평균)": {
           'type': 'ratio',
           'field_name': '무형자산/총자산',
           'condition_type': 'gt',
           'default_value': 0,
           'default_include': False
       },
       "유형자산/총자산(평균)": {
           'type': 'ratio',
           'field_name': '유형자산/총자산',
           'condition_type': 'gt',
           'default_value': 0,
           'default_include': False
       },
       "재고자산/총자산(평균)": {
           'type': 'ratio',
           'field_name': '재고자산/총자산',
           'condition_type': 'gt',
           'default_value': 0,
           'default_include': False
       },
       "재고자산보유일수(365/재고자산회전율)(평균)": {
           'type': 'ratio',
           'field_name': '재고자산보유일수',
           'condition_type': 'gt',
           'default_value': 0,
           'default_include': False
       }
   }
   # X비교 매핑
   COMPARISON_MAPPING = {
       "초과": "gt",
       "이상": "gte",
       "미만": "lt",
       "이하": "lte",
       "같음": "eq",
       "공란": "blank",
       "공란아님": "not_blank",
       "텍스트 일치": "equals",
       "텍스트 포함": "contains",
       "All equals": "all_equals"
   }
   # 포함/제외 매핑
   INCLUDE_MAPPING = {
       "포함": True,
       "제외": False
   }
   def __init__(self, start_year, end_year):
       """
       Parameters:
       - start_year: 분석 시작 연도
       - end_year: 분석 종료 연도
       """
       self.start_year = start_year
       self.end_year = end_year
       self.num_years = end_year - start_year + 1
   def convert_simple_input(self, user_criteria):
       """
       단순화된 사용자 입력을 프로그램 config로 변환
       Parameters:
       - user_criteria: 딕셔너리 또는 리스트
         {
             'account': 'Account 이름',
             'xValue': 숫자 또는 텍스트 (선택),
             'xCompare': '초과' | '이상' | ... (선택),
             'include': '포함' | '제외' (선택)
         }
       Returns:
       - config 딕셔너리 리스트
       """
       if isinstance(user_criteria, list):
           return [self._convert_single_simple_criteria(c) for c in user_criteria]
       else:
           return self._convert_single_simple_criteria(user_criteria)
       
   def _convert_single_simple_criteria(self, user_input):
       """단일 기준 변환 (단순화)"""
       # NaN 값 처리
       def safe_str(value, default=''):
           if pd.isna(value):
               return default
           return str(value).strip()
       def safe_value(value, default=None):
           if pd.isna(value):
               return default
           return value
       # 사용자 입력 파싱
       account = safe_str(user_input.get('account', ''))
       user_x_value = safe_value(user_input.get('xValue'))
       user_x_compare = safe_str(user_input.get('xCompare', ''))
       user_include = safe_str(user_input.get('include', ''))
       # Account 설정 가져오기
       if account not in self.ACCOUNT_CONFIG:
           print(f"Warning: 알 수 없는 Account '{account}'. 건너뜁니다.")
           return None
       account_config = self.ACCOUNT_CONFIG[account].copy()
       # xValue 오버라이드 (사용자가 입력한 경우)
       if user_x_value is not None:
           account_config['value'] = user_x_value
       else:
           account_config['value'] = account_config.get('default_value', 0)
       # xCompare 오버라이드 (사용자가 입력한 경우)
       if user_x_compare:
           mapped_compare = self.COMPARISON_MAPPING.get(user_x_compare)
           if mapped_compare:
               account_config['condition_type'] = mapped_compare
       # include 오버라이드 (사용자가 입력한 경우)
       if user_include:
           mapped_include = self.INCLUDE_MAPPING.get(user_include)
           if mapped_include is not None:
               account_config['include'] = mapped_include
       else:
           account_config['include'] = account_config.get('default_include', True)
       # 불필요한 default_ 키 제거
       account_config.pop('default_value', None)
       account_config.pop('default_include', None)
       return account_config
   

   def load_criteria_from_excel(self, excel_path, sheet_name="컨트롤시트"):
       """
       엑셀 파일에서 단순화된 기준 정보를 읽어옴
       Parameters:
       - excel_path: 엑셀 파일 경로
       - sheet_name: 시트 이름 (기본값: "컨트롤시트")
       Returns:
       - 변환된 criteria_configs 리스트
       """
       try:
           df = pd.read_excel(excel_path, sheet_name=sheet_name)
       except Exception as e:
           print(f"Error: 컨트롤시트를 읽는 중 오류 발생: {e}")
           return []
       # 필요한 컬럼 확인
       required_columns = ['account']
       optional_columns = ['xValue', 'xCompare', 'include']
       if 'account' not in df.columns:
           print(f"Error: 컨트롤시트에 'account' 컬럼이 없습니다.")
           return []
       # account가 비어있는 행 제거
       df = df.dropna(subset=['account'], how='all')
       # 사용 가능한 컬럼만 선택
       available_columns = [col for col in required_columns + optional_columns if col in df.columns]
       # 각 행을 딕셔너리로 변환
       user_criteria_list = df[available_columns].to_dict('records')
       # 프로그램 config로 변환
       converted_configs = []
       for criteria in user_criteria_list:
           config = self._convert_single_simple_criteria(criteria)
           if config is not None:  # None이 아닌 것만 추가
               converted_configs.append(config)
       print(f"✓ 컨트롤시트에서 {len(converted_configs)}개 기준을 읽어왔습니다.")
       return converted_configs


def main_processor(payload):
    
    input_data = payload["inputData"]
    criteria_list = payload["criteriaList"]

    # input_data = {
    #     "corpName": corp_name,
    #     "targetCorp": target_corp,
    #     "yearFrom": year_from,
    #     "yearTo": year_to,
    #     "rawFilePath": self.file_path
    # }
    #         criteria_list.append({
    #     "seq": idx,
    #     "account": account,
    #     "xValue": x_value,
    #     "xCompare": x_compare,
    #     "include": include
    # })
    
    converter = SimpleUserInputConverter(start_year=input_data["yearFrom"], end_year=input_data["yearTo"])
    converted = converter.convert_simple_input(criteria_list)

    processor = Analysis(
        tested_party=input_data["targetCorp"],
        name=input_data["corpName"],
        start_year=input_data["yearFrom"],
        end_year=input_data["yearTo"],
        number_of_criteria=len(criteria_list),
        data_path=input_data["rawFilePath"],
        criteria_list=criteria_list
    )

    # 포맷 생성
    processor.create_format()
    # Raw 데이터 채우기 (파일이 있다면)
    processor._populate_raw_data_from_excel()
    # 계산 수식 삽입
    processor.insert_formular()

    processor.apply_quantitative_criteria_formulas(converted)
    
    # 통과/탈락 요약 수식 (Row 20, 21)
    processor.insert_pass_fail_summary()

    # 최종 서식 적용 (Accounting, %, Border)
    processor.apply_final_styles()

    processor.save_file()

if __name__ == "__main__":
    test_payload = {'inputData': {'corpName': 'Test', 'targetCorp': 'Test', 'yearFrom': 2021, 'yearTo': 2023, 'rawFilePath': 'C:/Users/JX851XF/OneDrive - EY/Desktop/Python/BM/raw_1.xlsx'}, 'criteriaList': [{'seq': 1, 'account': '상장여부', 'xValue': 'Listed', 'xCompare': '텍스트 일치', 'include': True}, {'seq': 2, 'account': '영업이익(평균)', 'xValue': '0', 'xCompare': '미만', 'include': False}, {'seq': 3, 'account': '감사의견', 'xValue': 'Unqualified', 'xCompare': 'All equals', 'include': True}]}

    main_processor(test_payload)